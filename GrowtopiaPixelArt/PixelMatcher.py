"""
Growtopia像素画匹配器
把图片里的每个像素,匹配成颜色最接近的道具。
生成:方块网格、预览图、六宫格对比图、货架预览图,还有Excel表格。
new(带方差)和old(纯色差)两种算法都会跑一遍。
支持用 output/blacklist.json 屏蔽某些道具。
用法: python PixelMatcher.py <图片> <宽> <高>
例子: python PixelMatcher.py pixel_art.png 70 70
输出位置: ./output/pixelmatcher/<图片名>/
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path
from typing import Optional

import numpy as np
from PIL import Image, ImageDraw, ImageFont

from gtCommon import clean_filename, compute_color_stats, sprite_path

# ============================================
# 配置
# ============================================
BASE_DIR = Path(__file__).resolve().parent
SPRITES_DIR = BASE_DIR / "sprites"
SOLID_FILE = BASE_DIR / "output" / "solid_items.json"
STATS_FILE = BASE_DIR / "output" / "item_stats.json"
BLACKLIST_FILE = BASE_DIR / "output" / "blacklist.json"
OUTPUT_DIR = BASE_DIR / "output" / "pixelmatcher"
SHELF_SPRITE = SPRITES_DIR / "Display Shelf.png"
SHELF_CHARCOAL = SPRITES_DIR / "Display Shelf Charcoal.png"
SPRITE_SIZE = 15
POSTERIZE_COLORS = 0
VARIANCE_WEIGHT = 0.05

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================
# 解析命令行参数
# ============================================
def parse_args() -> tuple[str, Optional[int], Optional[int]]:
    if len(sys.argv) == 4:
        image_path = sys.argv[1]
        try:
            pw, ph = int(sys.argv[2]), int(sys.argv[3])
        except ValueError:
            print("Width and height must be numbers!")
            sys.exit(1)
    elif len(sys.argv) == 2:
        image_path = sys.argv[1]
        pw, ph = None, None
    else:
        # 没传参数,那就直接问用户要
        image_path = input("Image path: ").strip().strip('"')
        dims = input("Pixel dimensions (width height, blank for auto): ").strip()
        if dims:
            try:
                pw, ph = map(int, dims.split())
            except ValueError:
                print("Invalid! Use: width height")
                sys.exit(1)
        else:
            pw, ph = None, None

    if not Path(image_path).exists():
        print(f"File not found: {image_path}")
        sys.exit(1)

    return image_path, pw, ph


# ============================================
# 小工具函数
# ============================================
def get_folder(image_path: str) -> Path:
    """根据图片名字,建好(或者拿到)专属的输出文件夹。"""
    base_name = Path(image_path).stem
    folder = OUTPUT_DIR / base_name
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# ============================================
# 黑名单
# ============================================
def load_blacklist() -> tuple[list[str], list[str]]:
    """从 blacklist.json 里读出被屏蔽的关键词和道具名单。"""
    if not BLACKLIST_FILE.exists():
        print("  No blacklist.json found, using all items")
        return [], []
    with open(BLACKLIST_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    patterns = data.get("banned_patterns", [])
    items = data.get("banned_items", [])
    print(f"  Blacklist loaded: {len(patterns)} patterns, {len(items)} items")
    return patterns, items


def is_banned(name: str, patterns: list[str], items: list[str]) -> bool:
    """看看这个道具名字是不是踩中了黑名单关键词,或者本来就在禁用名单里。"""
    name_lower = name.lower()
    return name in items or any(p in name_lower for p in patterns)


# ============================================
# 读取数据
# ============================================
def load_solid_items() -> list[str]:
    if not SOLID_FILE.exists():
        print("solid_items.json not found! Run SolidItemFilter.py first.")
        sys.exit(1)
    with open(SOLID_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def load_item_database(solid_items: list[str]) -> tuple[list[str], np.ndarray, np.ndarray]:
    """
    把道具颜色数据整理成三个对得上号的数组(名字、平均颜色、方差),
    这样后面匹配的时候可以用numpy一次性算完,不用挨个循环。

    有缓存(item_stats.json)就优先用缓存;缓存里没有的道具,
    就用gtCommon里那套跟ComputeItemStats.py一模一样的逻辑现算,
    这样两个脚本算出来的颜色永远不会对不上。
    """
    cached = {}
    if STATS_FILE.exists():
        print("Loading cached color stats...")
        with open(STATS_FILE, "r", encoding="utf-8") as f:
            cached = json.load(f)
    else:
        print("No item_stats.json cache found - computing all stats now "
              "(tip: run ComputeItemStats.py first to speed this up next time).")

    names: list[str] = []
    avgs: list[list[float]] = []
    variances: list[float] = []
    computed_fresh = 0  # 记一下有多少个道具是现算的,不是从缓存拿的

    for name in solid_items:
        entry = cached.get(name)

        if entry is None:
            path = sprite_path(SPRITES_DIR, name)
            if not path.exists():
                continue
            try:
                entry = compute_color_stats(path)
            except Exception as e:
                print(f"  Warning: failed to read sprite for '{name}': {e}")
                continue
            if entry is None:
                continue
            computed_fresh += 1

        names.append(name)
        avgs.append(entry["avg"])
        variances.append(entry["variance"])

    if computed_fresh:
        print(f"  {computed_fresh} items weren't in the cache and were computed on the fly.")

    print(f"  {len(names)} items ready")
    return names, np.array(avgs, dtype=np.float64), np.array(variances, dtype=np.float64)


# ============================================
# 匹配逻辑
# ============================================
def color_distances(pixel: tuple[int, int, int], avgs: np.ndarray) -> np.ndarray:
    """
    一次性算出某个像素跟所有道具平均颜色的加权色差。
    颜色越暗,色差会被放大得越多,这样阴影部分才不会全部糊成一个道具。
    """
    r, g, b = pixel
    dr = avgs[:, 0] - r
    dg = avgs[:, 1] - g
    db = avgs[:, 2] - b
    rgb_dist = np.sqrt(2 * dr**2 + 4 * dg**2 + 3 * db**2)

    pixel_light = (r + g + b) / 3
    item_light = avgs.mean(axis=1)
    avg_light = np.maximum((pixel_light + item_light) / 2, 10)  # 防止除以0
    amplification = 1.0 + (150 / (avg_light + 50))

    return rgb_dist * amplification


def match_pixels(
    image_path: str,
    pw: Optional[int],
    ph: Optional[int],
    names: list[str],
    avgs: np.ndarray,
    variances: np.ndarray,
):
    print(f"\nLoading: {Path(image_path).name}")
    img = Image.open(image_path).convert("RGB")
    orig_w, orig_h = img.size

    if pw and ph:
        print(f"  Original: {orig_w}x{orig_h} -> {pw}x{ph}")
    else:
        pw, ph = orig_w, orig_h
        print(f"  Size: {pw}x{ph} (auto)")

    if POSTERIZE_COLORS > 0:
        print(f"  Posterizing to {POSTERIZE_COLORS} colors...")
        img = img.quantize(colors=POSTERIZE_COLORS).convert("RGB")

    img = img.resize((pw, ph), Image.NEAREST)
    print(f"  Pixels: {pw * ph}")
    pixels = img.load()

    # 每种颜色只算一次色差,new和old两种算法共用这一份结果就够了,
    # 不用像以前那样为了两种算法各算一遍。
    cache: dict[tuple[int, int, int], tuple[str, str]] = {}
    grid_new, grid_old = [], []

    print("Matching (new + old algorithm)...")
    for y in range(ph):
        row_new, row_old = [], []
        for x in range(pw):
            px = pixels[x, y]

            if px not in cache:
                dist = color_distances(px, avgs)
                idx_old = int(np.argmin(dist))
                score = dist + variances * VARIANCE_WEIGHT
                idx_new = int(np.argmin(score))
                cache[px] = (names[idx_new], names[idx_old])

            name_new, name_old = cache[px]
            row_new.append(name_new)
            row_old.append(name_old)

        grid_new.append(row_new)
        grid_old.append(row_old)
        if ph > 10 and y % max(1, ph // 10) == 0:
            print(f"  Row {y + 1}/{ph}")

    return grid_new, grid_old, pw, ph


# ============================================
# 保存结果
# ============================================
def save_results(grid, w: int, h: int, image_path: str, suffix: str = "") -> Counter:
    folder = get_folder(image_path)
    base_name = Path(image_path).stem

    filename = f"{base_name}_grid{suffix}.json"
    grid_path = folder / filename
    with open(grid_path, "w", encoding="utf-8") as f:
        json.dump(grid, f, indent=2)

    counts = Counter()
    for row in grid:
        counts.update(row)

    algo_name = "NEW" if suffix == "" else "OLD"
    print(f"\n{'=' * 50}")
    print(f"RESULTS ({base_name}) - {algo_name} ALGORITHM")
    print(f"{'=' * 50}")
    print(f"  {w}x{h} | {w * h} pixels | {len(counts)} items")
    print(f"  Grid: {grid_path}")
    print(f"\n  {'Qty':>6}  Item")
    print(f"  {'-' * 40}")
    for item, qty in counts.most_common():
        print(f"  {qty:>6}x  {item}")

    return counts


# ============================================
# 拼货架图(new/old共用)
# ============================================
def build_shelf_image(grid, w: int, h: int, shelf_sprite_path: Path) -> Image.Image:
    ITEM_SIZE = 70
    SLOT_GAP = 5
    SHELF_SIZE = 192
    LEFT_PAD = 26
    TOP_PAD = 20
    BOTTOM_EXTRA = 15

    sprites: dict[str, Image.Image] = {}
    for row in grid:
        for name in row:
            if name not in sprites:
                path = sprite_path(SPRITES_DIR, name)
                if path.exists():
                    sprites[name] = Image.open(path).convert("RGBA").resize(
                        (ITEM_SIZE, ITEM_SIZE), Image.NEAREST
                    )
                else:
                    # 贴图丢了就用洋红色占位,方便一眼看出哪里缺图
                    sprites[name] = Image.new("RGBA", (ITEM_SIZE, ITEM_SIZE), (255, 0, 255, 255))

    shelf_bg = Image.open(shelf_sprite_path).convert("RGBA")
    if shelf_bg.size != (SHELF_SIZE, SHELF_SIZE):
        shelf_bg = shelf_bg.resize((SHELF_SIZE, SHELF_SIZE), Image.NEAREST)

    sw, sh = (w + 1) // 2, (h + 1) // 2
    total_w = sw * SHELF_SIZE
    total_h = sh * SHELF_SIZE

    img = Image.new("RGBA", (total_w, total_h), (50, 50, 50, 255))

    for sy in range(sh):
        for sx in range(sw):
            shelf = shelf_bg.copy()
            positions = [(sy * 2, sx * 2), (sy * 2, sx * 2 + 1), (sy * 2 + 1, sx * 2), (sy * 2 + 1, sx * 2 + 1)]
            left_x = LEFT_PAD
            right_x = LEFT_PAD + ITEM_SIZE + SLOT_GAP
            top_y = TOP_PAD
            bottom_y = TOP_PAD + ITEM_SIZE + SLOT_GAP + BOTTOM_EXTRA
            spots = [(left_x, top_y), (right_x, top_y), (left_x, bottom_y), (right_x, bottom_y)]

            for (gy, gx), (px, py) in zip(positions, spots):
                if gy < h and gx < w:
                    shelf.paste(sprites[grid[gy][gx]], (px, py), sprites[grid[gy][gx]])

            img.paste(shelf, (sx * SHELF_SIZE, sy * SHELF_SIZE))

    return img


# ============================================
# 预览图 + 六宫格对比图
# ============================================
def build_preview(grid_new, grid_old, w: int, h: int, image_path: str) -> None:
    folder = get_folder(image_path)
    base_name = Path(image_path).stem

    sprite_cache: dict[tuple[str, int], Image.Image] = {}
    print("\nBuilding previews...")

    def get_sprite(name: str, size: int) -> Image.Image:
        key = (name, size)
        if key not in sprite_cache:
            path = sprite_path(SPRITES_DIR, name)
            if path.exists():
                sprite_cache[key] = Image.open(path).convert("RGBA").resize((size, size), Image.NEAREST)
            else:
                sprite_cache[key] = Image.new("RGBA", (size, size), (255, 0, 255, 255))
        return sprite_cache[key]

    def build_flat_preview(grid) -> Image.Image:
        """把网格里每个格子的道具贴图拼成一张平铺的大图。"""
        preview = Image.new("RGBA", (w * SPRITE_SIZE, h * SPRITE_SIZE))
        for y in range(h):
            for x in range(w):
                preview.paste(get_sprite(grid[y][x], SPRITE_SIZE), (x * SPRITE_SIZE, y * SPRITE_SIZE))
        return preview

    pw, ph = w * SPRITE_SIZE, h * SPRITE_SIZE

    original = Image.open(image_path).convert("RGBA").resize((pw, ph), Image.NEAREST)
    flat_new = build_flat_preview(grid_new)
    flat_old = build_flat_preview(grid_old)

    flat_new.save(folder / f"{base_name}_preview.png")
    flat_old.save(folder / f"{base_name}_preview_old.png")
    print(f"  Preview (new): {folder}/{base_name}_preview.png")
    print(f"  Preview (old): {folder}/{base_name}_preview_old.png")

    shelf_new = build_shelf_image(grid_new, w, h, SHELF_SPRITE)
    shelf_new_rgb = shelf_new.convert("RGBA").resize((pw, ph), Image.NEAREST)
    shelf_new.save(folder / f"{base_name}_shelves.png")
    print(f"  Shelf (new): {folder}/{base_name}_shelves.png")

    if SHELF_CHARCOAL.exists():
        shelf_charcoal = build_shelf_image(grid_new, w, h, SHELF_CHARCOAL)
        shelf_charcoal_rgb = shelf_charcoal.convert("RGBA").resize((pw, ph), Image.NEAREST)
        shelf_charcoal.save(folder / f"{base_name}_shelves_charcoal.png")
        print(f"  Shelf (charcoal): {folder}/{base_name}_shelves_charcoal.png")
    else:
        # 没有炭灰色货架贴图的话,就拿一张纯色背景顶替一下
        shelf_charcoal_rgb = Image.new("RGBA", (pw, ph), (50, 50, 50, 255))

    shelf_old = build_shelf_image(grid_old, w, h, SHELF_SPRITE)
    shelf_old_rgb = shelf_old.convert("RGBA").resize((pw, ph), Image.NEAREST)
    shelf_old.save(folder / f"{base_name}_shelves_old.png")
    print(f"  Shelf (old): {folder}/{base_name}_shelves_old.png")

    gap = 4
    total_w = pw * 3 + gap * 2
    total_h = ph * 2 + gap
    comp = Image.new("RGBA", (total_w, total_h), (30, 30, 30, 255))

    comp.paste(original, (0, 0))
    comp.paste(flat_new, (pw + gap, 0))
    comp.paste(flat_old, (pw * 2 + gap * 2, 0))

    comp.paste(shelf_new_rgb, (0, ph + gap))
    comp.paste(shelf_charcoal_rgb, (pw + gap, ph + gap))
    comp.paste(shelf_old_rgb, (pw * 2 + gap * 2, ph + gap))

    draw = ImageDraw.Draw(comp)
    try:
        font = ImageFont.truetype("arial.ttf", 16)
    except OSError:
        # 找不到arial字体就用默认字体凑合
        font = ImageFont.load_default()

    labels = [
        ("ORIGINAL", (5, 5)),
        ("ITEMS (NEW)", (pw + gap + 5, 5)),
        ("ITEMS (OLD)", (pw * 2 + gap * 2 + 5, 5)),
        ("SHELF (NEW)", (5, ph + gap + 5)),
        ("SHELF CHARCOAL", (pw + gap + 5, ph + gap + 5)),
        ("SHELF (OLD)", (pw * 2 + gap * 2 + 5, ph + gap + 5)),
    ]
    for text, pos in labels:
        draw.text(pos, text, fill=(255, 255, 0), font=font)

    comparison_path = folder / f"{base_name}_comparison.png"
    comp.save(comparison_path)
    print(f"  Comparison (6-panel): {comparison_path}")


# ============================================
# 导出Excel
# ============================================
def export_excel(grid_new, grid_old, w: int, h: int, image_path: str) -> None:
    folder = get_folder(image_path)
    base_name = Path(image_path).stem
    print("\nExporting Excel...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed! pip install openpyxl")
        return

    label_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    def write_grid_sheet(ws, grid) -> None:
        """把方块网格写进一个Excel工作表,顺便标好行号列号。"""
        for col in range(1, w + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        for y in range(h):
            ws.row_dimensions[y + 1].height = 20
            for x in range(w):
                c = ws.cell(row=y + 1, column=x + 1, value=grid[y][x])
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.font = Font(name="Calibri", size=9)
        for i in range(h):
            c = ws.cell(row=i + 1, column=w + 2, value=f"Row {i + 1}")
            c.fill, c.font = label_fill, label_font
        for i in range(w):
            c = ws.cell(row=h + 1, column=i + 1, value=f"Col {i + 1}")
            c.fill, c.font = label_fill, label_font

    def write_shelf_layout_sheet(ws, grid) -> None:
        """把每4个格子拼成一个货架,写进货架布局工作表。"""
        sw, sh = (w + 1) // 2, (h + 1) // 2
        for col in range(1, sw + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        for sy in range(sh):
            ws.row_dimensions[sy + 1].height = 40
            for sx in range(sw):
                items = []
                for dy, dx in [(0, 0), (0, 1), (1, 0), (1, 1)]:
                    gy, gx = sy * 2 + dy, sx * 2 + dx
                    if gy < h and gx < w:
                        items.append(grid[gy][gx])
                c = ws.cell(row=sy + 1, column=sx + 1, value="\n".join(items))
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.font = Font(name="Calibri", size=8)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Build Grid (New)"
    write_grid_sheet(ws1, grid_new)

    ws1b = wb.create_sheet("Build Grid (Old)")
    write_grid_sheet(ws1b, grid_old)

    counts_new = Counter(item for row in grid_new for item in row)
    counts_old = Counter(item for row in grid_old for item in row)

    ws2 = wb.create_sheet("Shopping List")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 5
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 12

    ws2.cell(row=1, column=1, value="Item (New)").font = Font(bold=True, color="4472C4")
    ws2.cell(row=1, column=2, value="Qty").font = Font(bold=True, color="4472C4")
    ws2.cell(row=1, column=4, value="Item (Old)").font = Font(bold=True, color="e94560")
    ws2.cell(row=1, column=5, value="Qty").font = Font(bold=True, color="e94560")

    sorted_new = counts_new.most_common()
    sorted_old = counts_old.most_common()
    max_rows = max(len(sorted_new), len(sorted_old))

    for i in range(max_rows):
        row = i + 2
        if i < len(sorted_new):
            ws2.cell(row=row, column=1, value=sorted_new[i][0])
            ws2.cell(row=row, column=2, value=sorted_new[i][1])
        if i < len(sorted_old):
            ws2.cell(row=row, column=4, value=sorted_old[i][0])
            ws2.cell(row=row, column=5, value=sorted_old[i][1])

    total_row = max_rows + 3
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=sum(counts_new.values())).font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=5, value=sum(counts_old.values())).font = Font(bold=True)

    ws3 = wb.create_sheet("Shelf Layout (New)")
    write_shelf_layout_sheet(ws3, grid_new)

    ws4 = wb.create_sheet("Shelf Layout (Old)")
    write_shelf_layout_sheet(ws4, grid_old)

    excel_path = folder / f"{base_name}_build.xlsx"
    wb.save(excel_path)
    print(f"  Saved: {excel_path}")
    print("  Sheets: Build Grid (New) | Build Grid (Old) | Shopping List | Shelf Layout (New) | Shelf Layout (Old)")


# ============================================
# 主程序
# ============================================
def main() -> None:
    print("=" * 50)
    print("GROWTOPIA PIXEL MATCHER")
    print("=" * 50)

    image_path, pw, ph = parse_args()
    solid_items = load_solid_items()

    # 先按黑名单过滤一遍
    banned_patterns, banned_items = load_blacklist()
    before = len(solid_items)
    solid_items = [s for s in solid_items if not is_banned(s, banned_patterns, banned_items)]
    print(f"  Items after blacklist: {len(solid_items)} (removed {before - len(solid_items)})\n")

    names, avgs, variances = load_item_database(solid_items)
    grid_new, grid_old, w, h = match_pixels(image_path, pw, ph, names, avgs, variances)

    save_results(grid_new, w, h, image_path)
    save_results(grid_old, w, h, image_path, "_old")

    build_preview(grid_new, grid_old, w, h, image_path)
    export_excel(grid_new, grid_old, w, h, image_path)

    folder = get_folder(image_path)
    print(f"\n{'=' * 50}")
    print("ALL DONE!")
    print(f"{'=' * 50}")
    print(f"  {folder}/")
    print("    - *_grid.json (new)")
    print("    - *_grid_old.json (old)")
    print("    - *_preview.png (new)")
    print("    - *_preview_old.png (old)")
    print("    - *_comparison.png (6-panel)")
    print("    - *_shelves.png (new)")
    print("    - *_shelves_charcoal.png")
    print("    - *_shelves_old.png")
    print("    - *_build.xlsx (5 sheets)")


if __name__ == "__main__":
    main()
